import os
import json
import requests
import time
import re
from datetime import datetime
from io import BytesIO
from dotenv import load_dotenv
import openpyxl # Biblioteca para ler Excel

# Carrega variáveis de ambiente
load_dotenv()

GOOGLE_MAPS_API_KEY = os.environ.get("GOOGLE_MAPS_API_KEY")

# URL PARA BAIXAR COMO EXCEL (XLSX)
SHEET_ID = "1THVJ8O_P19UkHq6DMgcfNF77fyD4lNWlmZA_rOM9FY4"
SHEET_XLSX_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

CACHE_FILE = 'latlon_cache.json'
OUTPUT_FILE = 'ensaios.json'

def load_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_cache(cache_data):
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache_data, f, ensure_ascii=False, indent=4)

def get_google_coords(local_text, cache):
    key = local_text.strip()
    if not key: return None, None, False
    if key in cache: return cache[key]['lat'], cache[key]['lon'], False
    if not GOOGLE_MAPS_API_KEY: return None, None, False

    search_query = f"{key}, Belo Horizonte, MG"
    url = f"https://maps.googleapis.com/maps/api/geocode/json?address={search_query}&key={GOOGLE_MAPS_API_KEY}"
    
    try:
        response = requests.get(url, timeout=5)
        data = response.json()
        if data['status'] == 'OK':
            loc = data['results'][0]['geometry']['location']
            cache[key] = {'lat': loc['lat'], 'lon': loc['lng']}
            return loc['lat'], loc['lng'], True
    except Exception as e:
        print(f"   [x] Erro API para '{key}': {e}")
    return None, None, False

def extract_hyperlink(cell):
    """
    Tenta extrair link de uma célula do Excel.
    Pode estar no atributo .hyperlink ou dentro de uma fórmula =HYPERLINK()
    """
    # 1. Hiperlink direto (Inserir > Link)
    if cell.hyperlink:
        return cell.hyperlink.target
    
    # 2. Fórmula =HYPERLINK("url", "texto")
    if cell.value and isinstance(cell.value, str) and str(cell.value).upper().startswith('=HYPERLINK'):
        match = re.search(r'"(http[^"]+)"', cell.value)
        if match:
            return match.group(1)
            
    return ""

def processar_ensaios():
    print(">>> 1. Baixando planilha como EXCEL (.xlsx)...")
    
    try:
        response = requests.get(SHEET_XLSX_URL)
        if response.status_code != 200:
            print(f"   [ERRO] Status Code: {response.status_code}")
            return
    except Exception as e:
        print(f"   [ERRO] Falha no download: {e}")
        return

    print(">>> 2. Lendo arquivo Excel em memória...")
    
    # Carrega o Excel da memória (sem salvar no disco)
    try:
        wb = openpyxl.load_workbook(filename=BytesIO(response.content), data_only=True)
        # data_only=True tenta pegar o valor calculado, mas remove formulas. 
        # Porém precisamos das formulas para links as vezes. 
        # Vamos usar o modo padrão e tratar manualmente.
        wb = openpyxl.load_workbook(filename=BytesIO(response.content))
        ws = wb.active # Pega a primeira aba
    except Exception as e:
        print(f"   [ERRO] Falha ao abrir Excel: {e}")
        return

    cache_geo = load_cache()
    ensaios_processados = []
    api_calls = 0
    dias_semana = {0: 'Seg', 1: 'Ter', 2: 'Qua', 3: 'Qui', 4: 'Sex', 5: 'Sáb', 6: 'Dom'}

    # Mapa de índices
    idx_map = {} 
    header_found = False

    # Itera sobre as linhas do Excel
    # values_only=False para podermos acessar o objeto cell e pegar o hyperlink
    for row in ws.iter_rows():
        # Pega o texto de cada célula (convertendo None para "")
        row_text = [str(cell.value).upper() if cell.value else "" for cell in row]
        
        # Detecção de cabeçalho
        keywords_found = 0
        if any("BLOCO" in t for t in row_text): keywords_found += 1
        if any("DATA" in t for t in row_text): keywords_found += 1
        
        if keywords_found >= 2:
            print("   - Cabeçalho detectado!")
            for i, text in enumerate(row_text):
                if "BLOCO" in text: idx_map['BLOCO'] = i
                elif "DATA" in text: idx_map['DATA'] = i
                elif "HORÁRIO" in text or "HORARIO" in text: idx_map['HORA'] = i
                elif "LOCAL" in text: idx_map['LOCAL'] = i
                elif "VALOR" in text: idx_map['VALOR'] = i
            header_found = True
            continue

        if not header_found: continue

        try:
            # Verifica tamanho da linha
            if len(row) <= max(idx_map.values(), default=0): continue

            # Extração de dados (cell.value)
            cell_bloco = row[idx_map['BLOCO']]
            nome = str(cell_bloco.value).strip() if cell_bloco.value else ""
            
            if not nome or "responsável" in nome.lower() or "VOU PRO BLOCO" in nome.upper(): continue

            cell_local = row[idx_map['LOCAL']]
            local_raw = str(cell_local.value).strip() if cell_local.value else ""
            
            cell_data = row[idx_map['DATA']]
            data_raw = str(cell_data.value).strip() if cell_data.value else ""
            
            cell_hora = row[idx_map['HORA']]
            hora_raw = str(cell_hora.value).strip() if cell_hora.value else ""
            
            # --- Extração Link (Coluna Valor) ---
            link_ingresso = ""
            if 'VALOR' in idx_map:
                cell_valor = row[idx_map['VALOR']]
                link_ingresso = extract_hyperlink(cell_valor)

        except Exception as e:
            # print(f"Erro linha: {e}")
            continue

        # --- Tratamento de Data ---
        dt_iso = None
        data_display = f"{data_raw} - {hora_raw}"
        
        try:
            # Excel as vezes retorna data como datetime object direto
            if isinstance(cell_data.value, datetime):
                dia = cell_data.value.day
                mes = cell_data.value.month
            else:
                dia, mes = map(int, data_raw.split('/'))
            
            ano = 2025 if mes > 6 else 2026
            
            # Limpeza Hora
            hora_clean = hora_raw.lower().replace('h', ':').replace('30:00', '30')
            if hora_clean.endswith(':'): hora_clean += "00"
            if ':' not in hora_clean and hora_clean.isdigit(): hora_clean += ":00"
            
            dt_obj = datetime(ano, mes, dia)
            try:
                parts = hora_clean.split(':')
                h = int(parts[0])
                m = int(parts[1]) if len(parts) > 1 and parts[1] else 0
                dt_obj = dt_obj.replace(hour=h, minute=m)
            except: pass

            dt_iso = dt_obj.isoformat()
            nome_dia = dias_semana[dt_obj.weekday()]
            hora_formatada = dt_obj.strftime('%H:%M')
            data_display = f"{dia:02d}/{mes:02d} ({nome_dia}) - {hora_formatada}"
        except: pass

        # --- Geolocalização ---
        lat, lon, used = get_google_coords(local_raw, cache_geo)
        if used:
            api_calls += 1
            save_cache(cache_geo)
            time.sleep(0.2)

        ensaios_processados.append({
            "id": str(hash(nome + data_display + "ensaio")),
            "titulo": nome,
            "endereco": local_raw,
            "local": "Belo Horizonte",
            "data": data_display,
            "dt_iso": dt_iso,
            "categoria": "Ensaio", 
            "categoria_display": "Ensaio",
            "descricao": "", 
            "link_ingresso": link_ingresso, 
            "tamanho": 2, 
            "lat": lat,
            "lon": lon,
            "is_ensaio": True,
            "is_kids": False,
            "is_lgbt": False,
            "is_pet": False,
            "status": "futuro"
        })

    print(f"\n>>> 3. Salvando {len(ensaios_processados)} ensaios em '{OUTPUT_FILE}'...")
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump({"eventos": ensaios_processados}, f, ensure_ascii=False, indent=4)
        
    print(f"   - Chamadas API Google: {api_calls}")

if __name__ == "__main__":
    processar_ensaios()